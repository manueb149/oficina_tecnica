import string, os
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.styles import Color, PatternFill, Font, Border, Fill, Alignment


def open_file_location():
    return os.path.dirname(os.path.abspath(__file__))+'/base.xlsx'

def save_file_location():
    return os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+'/static/download/brigada_reten.xlsx'

def writeRow(row, empleado, ws): 
    ws[string.ascii_uppercase[0]+str(row)]=empleado.area
    ws[string.ascii_uppercase[0]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws[string.ascii_uppercase[1]+str(row)]=empleado.nombres+' '+empleado.apellidos
    ws[string.ascii_uppercase[1]+str(row)].alignment = Alignment(horizontal='left', wrapText=True)
    ws[string.ascii_uppercase[2]+str(row)]=empleado.cargo
    ws[string.ascii_uppercase[2]+str(row)].alignment = Alignment(horizontal='left', wrapText=True)
    ws[string.ascii_uppercase[3]+str(row)]=empleado.no_empleado
    ws[string.ascii_uppercase[3]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws[string.ascii_uppercase[4]+str(row)]=empleado.no_flota
    ws[string.ascii_uppercase[4]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws[string.ascii_uppercase[5]+str(row)]=empleado.no_movil
    ws[string.ascii_uppercase[5]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws[string.ascii_uppercase[6]+str(row)]=empleado.vehiculo
    ws[string.ascii_uppercase[6]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    ws[string.ascii_uppercase[7]+str(row)]=empleado.ficha_vehiculo
    ws[string.ascii_uppercase[7]+str(row)].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

# print(os.path.abspath(__file__))
print(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))+'/static/download/brigada_reten.xlsx')
